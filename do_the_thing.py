import math
import os
import sys
from openpyxl import load_workbook
from openpyxl import worksheet
from openpyxl import workbook
os.chdir(sys.path[0])


starting_cell = ["C", "40"]


alphabet = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N" ,"O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
x_values = [1, 2, 4, 5, 7, 8, 1, 2, 4, 5, 7, 8, 1, 2, 4, 5, 7, 8, 1, 2, 4, 5, 7, 8, 1, 2, 4, 5, 7, 8, 1, 2, 4, 5, 7, 8, 1, 2, 4, 5, 7, 8, ]
y_values = [1, 1, 1, 1, 1, 1, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3, 3, 3, 4, 4, 4, 4, 4, 4, 5, 5, 5, 5, 5, 5, 6, 6, 6, 6, 6, 6, 7, 7, 7, 7, 7, 7, ]



used_teacher_angles =  []
used_whiteboard_angles = []

def get_distance_angle(x1, y1, x2, y2):
    # returns the distance between [x1; y1] and [x2; y2] (index 0) and angle adjacent to [x1; y1][x1; y2] (index 1)
    result = []
    dy = y1 - y2
    dx = x1 - x2
    result.append(math.sqrt((math.pow(dx, 2)) + (math.pow(dy, 2))))
    result.append(math.atan(dx/dy) / math.pi)
    return(result)

def get_window_distance_score(x2):
    score = math.pow(0.8, (x2 - 1))
    if score < 1:
        return(score)
    else:
        return(1)
    
def get_door_distance_score(x2, y2):
    distance = get_distance_angle(9, 0, x2, y2)[0]
    score = math.pow(0.8, (distance - 1))
    if score < 1:
        return(score)
    else:
        return(1)
    
def get_teacher_distance_score(x2, y2):
    global used_teacher_angles
    position = get_distance_angle(2, 0, x2, y2)
    colision_test_angle = math.trunc((math.trunc(position[1] * 75) / 75) * 100) / 100

    score = math.pow(1.085, (x2 - 1)) - 1

    if colision_test_angle in used_teacher_angles:
        score = 1
    else:
        used_teacher_angles.append(colision_test_angle)

    if score > 1:
        score = 1
    elif score < 0:
        score = 0
    return(score)

def get_whiteboard_distance_score(x2, y2):
    global used_whiteboard_angles
    position = get_distance_angle(4.5, 0, x2, y2)
    colision_test_angle = math.trunc((math.trunc(position[1] * 75) / 75) * 100) / 100

    score = math.pow(0.8, (x2 - 1))

    if colision_test_angle in used_whiteboard_angles:
        score = 0
    else:
        used_teacher_angles.append(colision_test_angle)

    if score > 1:
        score = 1
    elif score < 0:
        score = 0
    return(score)

def get_score(x2, y2):
    global weights
    final_score = math.trunc((25 * (get_teacher_distance_score(x2, y2) * weights[0] + get_whiteboard_distance_score(x2, y2) * weights[1] + get_window_distance_score(x2) * weights[2] + get_door_distance_score(x2, y2) * weights[3])) * 1000) / 1000
    return(final_score)

def file_to_num_array(input):
    #separates a .txt file into lines and converts it itno an array
    output = []
    char_index = 0
    stack = ""
    while char_index != len(input):
        if input[char_index] != "\n":
            stack += input[char_index]
            if char_index == len(input) - 1:
                output.append(float(stack))
        else:
            if len(stack) != 0:
                output.append((stack))
                stack = ""
            
        char_index += 1

    return(output)

weights_source = open("weights.txt", "r")
weights = file_to_num_array(weights_source.read())
weights.pop(0)
weights.pop(0)
index = 0
while index != len(weights):
    weights[index] = float(weights[index])
    index += 1
# 0 - teacher; 1 - whiteboard; 2 - window; 3 - door
weights_source.close()

active_column = starting_cell[0]
active_row = starting_cell[1]

print(weights)

def move_cell(dir):
    global active_column
    global active_row
    match dir:
        case "w":
            active_row = str(int(active_row) - 1)
        case "s":
            active_row = str(int(active_row) + 1)
        case "a":
            active_column = alphabet[alphabet.index(active_column) - 1]
        case "d":
            active_column = alphabet[alphabet.index(active_column) + 1]
    

def print_class(scores):
    global starting_cell
    global active_column
    global active_row

    gaps = [1, 3, 7, 9, 13, 15, 19, 21, 25, 27, 31, 33, 37, 39]
    ends = [5, 11, 17, 23, 29, 35, 41]

    wb = load_workbook("UI.xlsx")
    ws = wb["UI"]
    array_index = 0
    while array_index != len(scores) and array_index != 42:
        ws[active_column + active_row] = scores[array_index]
        if array_index in gaps:
            move_cell("d")
            move_cell("d")
        elif array_index in ends:
            active_column = starting_cell[0]
            move_cell("s")
        else:
            move_cell("d")

        array_index += 1

    wb.save("UI.xlsx")

scores = []
index = 0
while index != 42:
    scores.append(get_score(x_values[index], y_values[index]))
    index += 1

print_class(scores)

os.system("start EXCEL.EXE UI.xlsx")
#print_class([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41])






